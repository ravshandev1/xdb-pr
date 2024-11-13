from fastapi import APIRouter, Depends, Request, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse, FileResponse
from .schemas import RegisterSchema, ChangePasswordSchema, LoginSchema, UserSchema, ApplicationSchema, \
    ApplicationUpdateSchema, UserUpdateSchema, ApplicationPeriodQuery
from .models import User, Application
from utils import get_jwt_token, hashed_password, verify_password, JWTAuth, paginate
from tasks import send_data_to_tax_task
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date
from calendar import monthrange
from pytz import timezone
import os
from requests import post
from tasks import ENV

router = APIRouter()


@router.post("/register")
async def register_user(data: RegisterSchema, user: User = Depends(JWTAuth())):
    if user.superuser:
        if await User.exists(phone=data.phone):
            return JSONResponse({'success': False, 'message': 'User already exists'}, status_code=400)
        if data.password != data.confirm_password:
            return JSONResponse({'success': False, 'message': 'Passwords do not match'}, status_code=400)
        data.password = hashed_password(data.password)
        del data.confirm_password
        await User.create(**data.__dict__)
        return {"success": True}
    return JSONResponse({'success': False, 'message': 'Only admin creates a user'}, status_code=403)


@router.delete("/user/{pk}")
async def delete_user(pk: int, user: User = Depends(JWTAuth())):
    if user.superuser:
        user = await User.get_or_none(id=pk)
        if user.superuser:
            return JSONResponse({'success': False, 'message': 'Superuser can not delete'}, status_code=400)
        if user is None:
            return JSONResponse({'success': False, 'message': 'User does not exist'}, status_code=400)
        await user.delete()
        return {"success": True}
    return JSONResponse({'success': False, 'message': 'Only admin deletes a user'}, status_code=403)


@router.get("/users")
async def get_users(user: User = Depends(JWTAuth())):
    if user.superuser:
        ls = list()
        for i in await User.exclude(superuser=True):
            ls.append(UserSchema.to_dict(i))
        return ls
    return JSONResponse({'success': False, 'message': 'Only admin deletes a user'}, status_code=403)


@router.post("/login")
async def login_user(data: LoginSchema):
    user = await User.get_or_none(phone=data.phone)
    if not user:
        return JSONResponse({'message': 'Username not found'}, status_code=404)
    if not verify_password(data.password, user.password):
        return JSONResponse({'message': 'Password is incorrect'}, status_code=400)
    token = get_jwt_token(user.id)
    return {'access_token': token}


@router.get("/user")
async def get_user(user: User = Depends(JWTAuth())):
    return UserSchema.to_dict(user)


@router.patch("/change-password")
async def change_password(data: ChangePasswordSchema, user: User = Depends(JWTAuth())):
    if data.new_password != data.confirm_password:
        return JSONResponse({"message": "Passwords have not the same"}, status_code=400)
    if not verify_password(data.old_password, user.password):
        return JSONResponse({'message': 'Old password is incorrect'}, status_code=400)
    user.password = hashed_password(data.new_password)
    await user.save()
    return {'message': 'Passwords have changed successfully'}


@router.patch("/user/{pk}", dependencies=[Depends(JWTAuth())])
async def update_user(pk: int, data: UserUpdateSchema):
    user = await User.get_or_none(id=pk)
    if not user:
        return JSONResponse({'message': 'User not found'}, status_code=404)
    if data.password:
        data.password = hashed_password(data.password)
    data = {k: v for k, v in data.__dict__.items() if v is not None}
    await user.update_from_dict(data)
    await user.save()
    return UserSchema.to_dict(user)


@router.get("/applications/{year}/{month}")
async def get_applications_by_month(year: int, month: int):
    try:
        applications = await Application.filter(date__range=[date(year, month, 1),
                                                             date(year, month, monthrange(year, month)[1])]).all()
        if not applications:
            raise HTTPException(status_code=404, detail="Applications not found")
        file_path = 'template.xlsx'
        wb = load_workbook(filename=file_path)
        ws = wb.active
        current_row = 2
        for app in applications:
            row_data = [
                str(app.code),
                str(app.date),
                str(app.area),
                str(app.river),
                str(app.plot),
                str(app.address),
                str(app.stir),
                str(app.dsi),
                str(app.subject_name),
                str(app.count),
                str(app.diff_count),
                str(app.status)
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                source_cell = ws.cell(row=current_row - 1, column=col)
                cell.font = source_cell.font.copy()
                cell.alignment = source_cell.alignment.copy()
                cell.border = source_cell.border.copy()
                cell.fill = source_cell.fill.copy()
                cell.number_format = source_cell.number_format
            current_row += 1
        new_file_name = "applications.xlsx"
        new_file_path = os.path.join(os.getcwd(), new_file_name)
        wb.save(new_file_path)
        return FileResponse(path=new_file_path, filename=new_file_name,
                                      media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print(e)
        return FileResponse(path="template.xlsx", filename="application",
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/applications", dependencies=[Depends(JWTAuth())])
async def get_applications(request: Request, query: ApplicationPeriodQuery = Depends()):
    ls = list()
    if query.year and query.month:
        qs = await Application.filter(date__range=[date(query.year, query.month, 1),
                         date(query.year, query.month, monthrange(query.year, query.month)[1])]).order_by('-id')
    else:
        qs = await Application.all().order_by('-id')
    if query.stir:
        qs = await Application.filter(stir=query.stir).order_by('-id')
    for i in qs:
        ls.append(ApplicationSchema.to_dict(i))
    return paginate(ls, request.url.path, query.limit, query.offset)


@router.post("/applications", dependencies=[Depends(JWTAuth())])
async def root(file: UploadFile = File(...)):
    if file.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        return JSONResponse({"success": False, "message": "Invalid file type. Please upload an Excel file."},
                                      status_code=400)
    content = await file.read()
    workbook = load_workbook(filename=BytesIO(content))
    sheet = workbook.active
    ls = list()
    ids = list()
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict()
            dt = date(int(row[1].split('.')[2]), int(row[1].split('.')[1]), int(row[1].split('.')[0]))
            if await Application.filter(date=dt, stir=row[6], dsi=row[7]).exists():
                await Application.create(code=row[0], date=dt, area=row[2], river=row[3], plot=row[4], address=row[5],
                                         stir=row[6], dsi=row[7], subject_name=row[8], count=row[9],
                                         status="Takrorlangan")
            else:
                app = await Application.create(code=row[0], date=dt, area=row[2], river=row[3], plot=row[4], dsi=row[7],
                                               address=row[5], stir=row[6], subject_name=row[8], count=row[9])
                data['send_id'] = row[0]
                data['send_date'] = f"{row[1]} {datetime.now().astimezone(timezone('Asia/Tashkent')).strftime('%H:%M:%S')}"
                data['ns10'] = row[7].split('/')[0]
                data['ns11'] = row[7].split('/')[1]
                data['address'] = row[5]
                data['tin'] = row[6]
                data['count'] = row[9]
                ls.append(data)
                ids.append({"id": app.id, "year": dt.year, "month": dt.month})
        send_data_to_tax_task.delay(ls, ids)
    except Exception as e:
        print(e)
    return {'success': True}


@router.patch("/application/{pk}")
async def update_application(pk: int, data: ApplicationUpdateSchema):
    obj = await Application.get_or_none(id=pk)
    if not obj:
        return JSONResponse({"message": "Application not found"}, status_code=404)
    if data.status:
        obj.status = data.status
    if data.diff_count:
        obj.diff_count = data.diff_count
    await obj.save()
    return {"success": True}


@router.get("/application/{pk}")
async def get_application(pk: int):
    obj = await Application.get_or_none(id=pk)
    if not obj:
        return JSONResponse({"message": "Application not found"}, status_code=404)
    data = dict()
    data['send_id'] = obj.code
    data['send_date'] = f"{obj.date} {datetime.now().astimezone(timezone('Asia/Tashkent')).strftime('%H:%M:%S')}"
    data['ns10'] = obj.dsi.split('/')[0]
    data['ns11'] = obj.dsi.split('/')[1]
    data['address'] = obj.address
    data['tin'] = obj.stir
    data['count'] = obj.count
    res = post(f"{ENV['TAX_API']}", json=data, headers={'Content-Type': 'application/json'})
    if res.status_code >= 400:
        obj.status = "Soliqni API si ishlamadi!"
    # elif res.status_code >= 400:
    #     obj.status = res.json()["text"]
    else:
        obj.status = "Muvofiqiyatli"
    await obj.save()
    return {"success": True}


@router.delete("/application/{pk}")
async def delete_application(pk: int):
    obj = await Application.get_or_none(id=pk)
    if not obj:
        return JSONResponse({"message": "Application not found"}, status_code=404)
    await obj.delete()
    return {"success": True}
